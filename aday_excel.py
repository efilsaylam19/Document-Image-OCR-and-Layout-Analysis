"""
Aday havuzunu Excel dosyasına yazar / günceller.
Gereksinim: pip install openpyxl
"""

import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_VAR = True
except ImportError:
    OPENPYXL_VAR = False

EXCEL_DOSYA = "aday_havuzu.xlsx"

SUTUNLAR = [
    ("ID",              8),
    ("Ad Soyad",        22),
    ("E-posta",         28),
    ("Telefon",         16),
    ("Sehir",           14),
    ("LinkedIn",        30),
    ("GitHub",          26),
    ("Egitim",          35),
    ("Deneyim (yil)",   14),
    ("Deneyim Ozet",    40),
    ("Beceriler",       45),
    ("Diller",          25),
    ("Kaynak Dosya",    30),
    ("Eklenme Tarihi",  18),
    ("Notlar",          35),
]

# ─── Renk paleti ───
BASLIK_DOLGU  = "1e3a5f"
BASLIK_YAZI   = "FFFFFF"
TEK_SATIR     = "e8f0fb"
CIFT_SATIR    = "FFFFFF"
KENAR_RENK    = "b0bec5"


def _kenarlık():
    ince = Side(style="thin", color=KENAR_RENK)
    return Border(left=ince, right=ince, top=ince, bottom=ince)


def _baslik_stili():
    return {
        "font":      Font(bold=True, color=BASLIK_YAZI, size=11, name="Segoe UI"),
        "fill":      PatternFill("solid", fgColor=BASLIK_DOLGU),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border":    _kenarlık(),
    }


def _veri_stili(satir_no: int, orta: bool = False):
    dolgu_renk = TEK_SATIR if satir_no % 2 == 0 else CIFT_SATIR
    return {
        "font":      Font(size=10, name="Segoe UI"),
        "fill":      PatternFill("solid", fgColor=dolgu_renk),
        "alignment": Alignment(
            horizontal="center" if orta else "left",
            vertical="center", wrap_text=True
        ),
        "border":    _kenarlık(),
    }


def _stil_uygula(hucre, stiller: dict):
    for k, v in stiller.items():
        setattr(hucre, k, v)


def excel_olustur_ya_ac(dosya_yolu: str):
    """Dosya yoksa oluştur, varsa aç. Çalışma sayfasını döndür."""
    if not OPENPYXL_VAR:
        raise ImportError("openpyxl kurulu değil: pip install openpyxl")

    if os.path.exists(dosya_yolu):
        wb = openpyxl.load_workbook(dosya_yolu)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Aday Havuzu"

        # Üst başlık
        ws.merge_cells("A1:O1")
        baslik_h = ws["A1"]
        baslik_h.value = "Aday Havuzu  —  Belge OCR & Layout Analiz Sistemi"
        _stil_uygula(baslik_h, {
            "font":      Font(bold=True, color=BASLIK_YAZI, size=14, name="Segoe UI"),
            "fill":      PatternFill("solid", fgColor="0d2137"),
            "alignment": Alignment(horizontal="center", vertical="center"),
        })
        ws.row_dimensions[1].height = 32

        # Sütun başlıkları (2. satır)
        for col_idx, (ad, gen) in enumerate(SUTUNLAR, start=1):
            h = ws.cell(row=2, column=col_idx, value=ad)
            _stil_uygula(h, _baslik_stili())
            ws.column_dimensions[get_column_letter(col_idx)].width = gen
        ws.row_dimensions[2].height = 26

        # Dondurulan satır
        ws.freeze_panes = "A3"

        wb.save(dosya_yolu)
        wb = openpyxl.load_workbook(dosya_yolu)
        ws = wb.active

    return wb, ws


def aday_satiri_yaz(ws, aday: dict, satir_no: int):
    """Tek bir adayı verilen satıra yazar."""
    deger_listesi = [
        aday.get("id", ""),
        aday.get("ad_soyad", ""),
        aday.get("email", ""),
        aday.get("telefon", ""),
        aday.get("sehir", ""),
        aday.get("linkedin", ""),
        aday.get("github", ""),
        aday.get("egitim_ozet", ""),
        aday.get("deneyim_yil", ""),
        aday.get("deneyim_ozet", ""),
        aday.get("beceri_str", ""),
        aday.get("dil_str", ""),
        aday.get("kaynak_dosya", ""),
        aday.get("eklenme_tarihi", datetime.now().strftime("%Y-%m-%d %H:%M")),
        aday.get("notlar", ""),
    ]
    for col_idx, deger in enumerate(deger_listesi, start=1):
        orta = col_idx in (1, 4, 5, 9, 14)
        h = ws.cell(row=satir_no, column=col_idx, value=deger)
        _stil_uygula(h, _veri_stili(satir_no - 2, orta=orta))
    ws.row_dimensions[satir_no].height = 22


def aday_ekle_excel(aday: dict, dosya_yolu: str = EXCEL_DOSYA) -> int:
    """
    Adayı Excel dosyasına ekler.
    Döndürür: yazılan satır numarası
    """
    wb, ws = excel_olustur_ya_ac(dosya_yolu)

    # Son dolu satırı bul (başlık 1-2, veriler 3'ten başlar)
    son_satir = ws.max_row
    if son_satir < 2:
        son_satir = 2
    yeni_satir = son_satir + 1

    # Benzersiz ID üret
    aday["id"] = yeni_satir - 2  # 1'den başlayan sıra no
    aday_satiri_yaz(ws, aday, yeni_satir)

    wb.save(dosya_yolu)
    return aday["id"]


def excel_yenile(adaylar: list, dosya_yolu: str = EXCEL_DOSYA):
    """
    Tüm aday listesini sıfırdan Excel'e yazar.
    (Veritabanı güncellemelerinden sonra çağrılır.)
    """
    if not OPENPYXL_VAR:
        raise ImportError("openpyxl kurulu değil: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Aday Havuzu"

    # Üst başlık
    son_harf = get_column_letter(len(SUTUNLAR))
    ws.merge_cells(f"A1:{son_harf}1")
    h = ws["A1"]
    h.value = f"Aday Havuzu  —  Son güncelleme: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    _stil_uygula(h, {
        "font":      Font(bold=True, color=BASLIK_YAZI, size=13, name="Segoe UI"),
        "fill":      PatternFill("solid", fgColor="0d2137"),
        "alignment": Alignment(horizontal="center", vertical="center"),
    })
    ws.row_dimensions[1].height = 30

    for col_idx, (ad, gen) in enumerate(SUTUNLAR, start=1):
        hucre = ws.cell(row=2, column=col_idx, value=ad)
        _stil_uygula(hucre, _baslik_stili())
        ws.column_dimensions[get_column_letter(col_idx)].width = gen
    ws.row_dimensions[2].height = 26
    ws.freeze_panes = "A3"

    for i, aday in enumerate(adaylar):
        aday_satiri_yaz(ws, aday, i + 3)

    wb.save(dosya_yolu)
