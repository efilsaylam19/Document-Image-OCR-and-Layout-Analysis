"""
Candidate Excel Export — writes / updates the candidate pool to an Excel file.
Requirement: pip install openpyxl
"""

import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_VAR = True
except ImportError:
    OPENPYXL_VAR = False

EXCEL_FILE = "candidate_pool.xlsx"

COLUMNS = [
    ("ID",                  8),
    ("Full Name",          22),
    ("Email",              28),
    ("Phone",              16),
    ("City",               14),
    ("LinkedIn",           30),
    ("GitHub",             26),
    ("Education",          35),
    ("Experience (yrs)",   14),
    ("Experience Summary", 40),
    ("Skills",             45),
    ("Languages",          25),
    ("Source File",        30),
    ("Date Added",         18),
    ("Notes",              35),
]

# ─── Color palette ───
HEADER_FILL  = "1e3a5f"
HEADER_FONT  = "FFFFFF"
ODD_ROW      = "e8f0fb"
EVEN_ROW     = "FFFFFF"
BORDER_COLOR = "b0bec5"


def _border():
    thin = Side(style="thin", color=BORDER_COLOR)
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_style():
    return {
        "font":      Font(bold=True, color=HEADER_FONT, size=11, name="Segoe UI"),
        "fill":      PatternFill("solid", fgColor=HEADER_FILL),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border":    _border(),
    }


def _row_style(row_number: int, centered: bool = False):
    fill_color = ODD_ROW if row_number % 2 == 0 else EVEN_ROW
    return {
        "font":      Font(size=10, name="Segoe UI"),
        "fill":      PatternFill("solid", fgColor=fill_color),
        "alignment": Alignment(
            horizontal="center" if centered else "left",
            vertical="center", wrap_text=True
        ),
        "border":    _border(),
    }


def _apply_style(cell, styles: dict):
    for key, value in styles.items():
        setattr(cell, key, value)


def _open_or_create(file_path: str):
    """Open the workbook if it exists, or create a new one with headers."""
    if not OPENPYXL_VAR:
        raise ImportError("openpyxl is not installed: pip install openpyxl")

    if os.path.exists(file_path):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Candidate Pool"

        # Top title row
        ws.merge_cells("A1:O1")
        title_cell = ws["A1"]
        title_cell.value = "Candidate Pool  —  Document OCR & Layout Analysis System"
        _apply_style(title_cell, {
            "font":      Font(bold=True, color=HEADER_FONT, size=14, name="Segoe UI"),
            "fill":      PatternFill("solid", fgColor="0d2137"),
            "alignment": Alignment(horizontal="center", vertical="center"),
        })
        ws.row_dimensions[1].height = 32

        # Column headers (row 2)
        for col_idx, (label, width) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=2, column=col_idx, value=label)
            _apply_style(cell, _header_style())
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[2].height = 26

        # Freeze header rows
        ws.freeze_panes = "A3"

        wb.save(file_path)
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

    return wb, ws


def _write_candidate_row(ws, candidate: dict, row_number: int):
    """Write a single candidate to the given row."""
    values = [
        candidate.get("id",                  ""),
        candidate.get("full_name",           ""),
        candidate.get("email",               ""),
        candidate.get("phone",               ""),
        candidate.get("city",                ""),
        candidate.get("linkedin",            ""),
        candidate.get("github",              ""),
        candidate.get("education_summary",   ""),
        candidate.get("experience_years",    ""),
        candidate.get("experience_summary",  ""),
        candidate.get("skills_str",          ""),
        candidate.get("languages_str",       ""),
        candidate.get("source_file",         ""),
        candidate.get("date_added", datetime.now().strftime("%Y-%m-%d %H:%M")),
        candidate.get("notes",               ""),
    ]
    for col_idx, value in enumerate(values, start=1):
        centered = col_idx in (1, 4, 5, 9, 14)
        cell = ws.cell(row=row_number, column=col_idx, value=value)
        _apply_style(cell, _row_style(row_number - 2, centered=centered))
    ws.row_dimensions[row_number].height = 22


def add_candidate_to_excel(candidate: dict, file_path: str = EXCEL_FILE) -> int:
    """
    Append a candidate to the Excel file.
    Returns the row number that was written.
    """
    wb, ws = _open_or_create(file_path)

    # Find the last used row (headers on rows 1-2, data starts at row 3)
    last_row = ws.max_row
    if last_row < 2:
        last_row = 2
    new_row = last_row + 1

    # Auto-generate a sequential ID
    candidate["id"] = new_row - 2  # 1-based sequence number
    _write_candidate_row(ws, candidate, new_row)

    wb.save(file_path)
    return candidate["id"]


def refresh_excel(candidates: list, file_path: str = EXCEL_FILE):
    """
    Rewrite the entire candidate list to Excel from scratch.
    Called after database updates to keep the file in sync.
    """
    if not OPENPYXL_VAR:
        raise ImportError("openpyxl is not installed: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Candidate Pool"

    # Top title row
    last_col = get_column_letter(len(COLUMNS))
    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws["A1"]
    title_cell.value = f"Candidate Pool  —  Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    _apply_style(title_cell, {
        "font":      Font(bold=True, color=HEADER_FONT, size=13, name="Segoe UI"),
        "fill":      PatternFill("solid", fgColor="0d2137"),
        "alignment": Alignment(horizontal="center", vertical="center"),
    })
    ws.row_dimensions[1].height = 30

    for col_idx, (label, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=label)
        _apply_style(cell, _header_style())
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 26
    ws.freeze_panes = "A3"

    for i, candidate in enumerate(candidates):
        _write_candidate_row(ws, candidate, i + 3)

    wb.save(file_path)
